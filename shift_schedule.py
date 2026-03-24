import streamlit as st
import pandas as pd
import calendar

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

# ---------------- PAGE SETUP ----------------
st.set_page_config(layout="wide")
st.title("📅 Shift Schedule Generator Application SBD-ELE")

# ---------------- SESSION ----------------
if "employees" not in st.session_state:
    st.session_state.employees = []

if "schedule" not in st.session_state:
    st.session_state.schedule = None

# ---------------- MONTH & YEAR ----------------
col1, col2 = st.columns(2)

with col1:
    year = st.number_input("Year", 2020, 2100, 2026)

with col2:
    month = st.selectbox("Month", list(calendar.month_name)[1:])

month_num = list(calendar.month_name).index(month)
days_in_month = calendar.monthrange(year, month_num)[1]

# ---------------- ADD EMPLOYEE ----------------
st.subheader("👨‍🏭 Employee Management")

with st.form("emp_form", clear_on_submit=True):
    c1, c2 = st.columns(2)

    with c1:
        name = st.text_input("Employee Name")

    with c2:
        code = st.text_input("Employee ID")

    submit = st.form_submit_button("Add Employee")

    if submit:
        if name and code:
            emp = {"Name": name, "ID": code}

            if emp not in st.session_state.employees:
                st.session_state.employees.append(emp)
                st.success("✅ Employee Added Successfully")
            else:
                st.warning("Employee already exists")
        else:
            st.error("Enter Name & ID")

# ---------------- SHOW EMPLOYEES ----------------
st.write("### Employee List")

if st.session_state.employees:
    st.write(pd.DataFrame(st.session_state.employees))
else:
    st.info("No employees added yet")

# ---------------- GENERATE EMPTY SCHEDULE ----------------
st.subheader("📊 Generate Shift Schedule")

if st.button("Generate Schedule"):

    data = []

    for emp in st.session_state.employees:
        row = {
            "Employee Name": emp["Name"],
            "Employee ID": emp["ID"]
        }

        for d in range(1, days_in_month + 1):
            row[str(d)] = ""

        data.append(row)

    st.session_state.schedule = pd.DataFrame(data)

# ---------------- DISPLAY WITH DROPDOWN ----------------
if st.session_state.schedule is not None:

    st.markdown(f"# 📅 Shift Schedule - {month} {year}")

    # Dropdown for shift columns only
    column_config = {
        col: st.column_config.SelectboxColumn(
            options=["", "A", "B", "C", "G", "WO"],
            required=False
        )
        for col in st.session_state.schedule.columns
        if col not in ["Employee Name", "Employee ID"]
    }

    edited_df = st.data_editor(
        st.session_state.schedule,
        use_container_width=True,
        column_config={
            "Employee Name": st.column_config.TextColumn(disabled=True),
            "Employee ID": st.column_config.TextColumn(disabled=True),
            **column_config
        },
        num_rows="dynamic"
    )

    st.session_state.schedule = edited_df

# ---------------- EXPORT TO EXCEL ----------------
if st.button("Export to Excel"):

    if st.session_state.schedule is not None:

        wb = Workbook()
        ws = wb.active

        df = st.session_state.schedule
        total_cols = len(df.columns)

        # -------- TITLE --------
        title = f"📅 Shift Schedule - {month} {year}"

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = title
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.font = Font(size=14, bold=True)

        ws.append([])

        # -------- TABLE --------
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # -------- STYLE --------
        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        shift_colors = {
            "A": "ADD8E6",
            "B": "90EE90",
            "C": "FFD580",
            "G": "DDA0DD",
            "WO": "FF7F7F"
        }

        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=total_cols):
            for cell in row:

                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

                if cell.row == 3:
                    cell.font = Font(bold=True)

                if cell.value in shift_colors:
                    cell.fill = PatternFill(
                        start_color=shift_colors[cell.value],
                        end_color=shift_colors[cell.value],
                        fill_type="solid"
                    )

        # -------- AUTO WIDTH --------
        for col_num in range(1, total_cols + 1):

            col_letter = get_column_letter(col_num)
            max_length = 0

            for row in ws.iter_rows(min_row=3, max_row=ws.max_row,
                                    min_col=col_num, max_col=col_num):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[col_letter].width = max_length + 2

        wb.save("shift_schedule.xlsx")

        st.success("✅ Excel Exported Successfully!")

    else:
        st.warning("Generate schedule first")